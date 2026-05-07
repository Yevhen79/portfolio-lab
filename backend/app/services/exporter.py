"""PDF / Excel export of portfolios."""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def export_excel(portfolio: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    bold = Font(bold=True, color="FFFFFF", name="Calibri")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    metric_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    metric_font = Font(bold=True, color="00D4FF")

    def write_header(row: int, title: str) -> int:
        ws.cell(row=row, column=1, value=title).font = bold
        ws.cell(row=row, column=1).fill = header_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        return row + 1

    row = 1
    row = write_header(row, "Portfolio Lab — Markowitz Portfolio Report")
    ws.cell(row=row, column=1, value="Name").font = metric_font
    ws.cell(row=row, column=2, value=portfolio.name)
    row += 1
    ws.cell(row=row, column=1, value="Type").font = metric_font
    ws.cell(row=row, column=2, value=portfolio.portfolio_type)
    row += 1
    ws.cell(row=row, column=1, value="Initial Capital").font = metric_font
    ws.cell(row=row, column=2, value=f"${portfolio.initial_capital:,.2f}")
    row += 1
    ws.cell(row=row, column=1, value="Created").font = metric_font
    ws.cell(row=row, column=2, value=str(portfolio.created_at))
    row += 2

    row = write_header(row, "Portfolio Metrics (annualized)")
    metrics = [
        ("Expected Return", f"{portfolio.expected_return_annual*100:.2f}%"),
        ("Volatility", f"{portfolio.volatility_annual*100:.2f}%"),
        ("Sharpe Ratio", f"{portfolio.sharpe_ratio:.3f}"),
        ("Sortino Ratio", f"{portfolio.sortino_ratio:.3f}"),
        ("VaR 95%", f"{portfolio.var_95_annual*100:.2f}%"),
        ("CVaR 95%", f"{portfolio.cvar_95_annual*100:.2f}%"),
        ("Historical Max DD", f"{portfolio.max_drawdown_estimate*100:.2f}%"),
        ("Risk-Free Rate", f"{portfolio.risk_free_rate*100:.2f}%"),
    ]
    for k, v in metrics:
        ws.cell(row=row, column=1, value=k).font = metric_font
        ws.cell(row=row, column=2, value=v)
        ws.cell(row=row, column=1).fill = metric_fill
        row += 1
    row += 1

    row = write_header(row, "Asset Allocation")
    headers = ["Symbol", "Name", "Category", "Weight %", "USD Amount"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    row += 1
    weights = portfolio.weights or []
    for w in weights:
        ws.cell(row=row, column=1, value=w["symbol"])
        ws.cell(row=row, column=2, value=w["name"])
        ws.cell(row=row, column=3, value=w["category"])
        ws.cell(row=row, column=4, value=f"{w['weight']*100:.2f}%")
        ws.cell(row=row, column=5, value=f"${w['amount_usd']:,.2f}")
        row += 1

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(portfolio: Any) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "title", parent=styles["Heading1"], fontSize=20,
        textColor=colors.HexColor("#00D4FF"),
    )
    h2 = ParagraphStyle(
        "h2", parent=styles["Heading2"], fontSize=14,
        textColor=colors.HexColor("#FF00AA"),
    )
    body = styles["BodyText"]

    story = []
    story.append(Paragraph("Portfolio Lab", title))
    story.append(Paragraph(f"<b>{portfolio.name}</b>", styles["Heading2"]))
    story.append(Paragraph(
        f"Portfolio type: <b>{portfolio.portfolio_type}</b> | "
        f"Created: {portfolio.created_at} | "
        f"Initial capital: <b>${portfolio.initial_capital:,.2f}</b>",
        body,
    ))
    story.append(Spacer(1, 0.5 * cm))

    # Metrics table
    story.append(Paragraph("Portfolio Metrics (annualized)", h2))
    metric_rows = [
        ["Expected Return", f"{portfolio.expected_return_annual*100:.2f}%"],
        ["Volatility", f"{portfolio.volatility_annual*100:.2f}%"],
        ["Sharpe Ratio", f"{portfolio.sharpe_ratio:.3f}"],
        ["Sortino Ratio", f"{portfolio.sortino_ratio:.3f}"],
        ["VaR 95%", f"{portfolio.var_95_annual*100:.2f}%"],
        ["CVaR 95%", f"{portfolio.cvar_95_annual*100:.2f}%"],
        ["Historical Max DD", f"{portfolio.max_drawdown_estimate*100:.2f}%"],
        ["Risk-Free Rate", f"{portfolio.risk_free_rate*100:.2f}%"],
    ]
    t = Table(metric_rows, colWidths=[7 * cm, 7 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#E2E8F0")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#1E293B")),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5 * cm))

    # Allocation
    story.append(Paragraph("Asset Allocation", h2))
    rows = [["Symbol", "Name", "Category", "Weight", "USD"]]
    for w in portfolio.weights or []:
        rows.append([
            w["symbol"], w["name"][:30], w["category"],
            f"{w['weight']*100:.2f}%", f"${w['amount_usd']:,.2f}",
        ])
    t = Table(rows, colWidths=[2.5 * cm, 6 * cm, 2.5 * cm, 2 * cm, 3 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#00D4FF")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
            [colors.HexColor("#FFFFFF"), colors.HexColor("#F1F5F9")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t)

    doc.build(story)
    return buf.getvalue()
